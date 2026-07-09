"""
generate_dog_dataset.py
-----------------------
Generates a realistic synthetic canine sensor dataset for 12 dogs,
all in the 10-27 kg weight range (vest-compatible medium dogs).
Each dog is monitored across three consecutive activity phases:
  REST (12 min) -> WALK (12 min) -> RUN (12 min)
Sampled at 1-minute intervals = 36 readings per dog, 432 rows total.

Physiological basis
-------------------
  Body temperature and respiratory rate are parameterised by body weight
  within the 10-27 kg range:
    - Lighter dogs (10 kg): rest body temp ~38.70 C, rest resp ~22 bpm
    - Heavier dogs (27 kg): rest body temp ~38.42 C, rest resp ~15 bpm
  Both metrics rise progressively during WALK and RUN phases.
  Heavier dogs have better thermoregulation and smaller temperature/resp
  rises for the same exercise intensity.

Labels follow the 2-of-3 Master Threshold Classification Table:
  body_temp    > 39.5 C      -> body stressed
  resp_rate   >= 29 bpm      -> resp stressed
  ambient_temp > 30.0 C      -> ambient stressed
  >= 2 of 3 stressed         -> label = 1  (HEAT_STRESSED)
  < 2 of 3 stressed          -> label = 0  (NORMAL)

Output files
------------
  canine_dog_dataset.csv   - training-ready CSV (use with train_model.py)
  canine_dog_dataset.xlsx  - color-coded Excel workbook with 3 sheets

Usage
-----
  pip install pandas numpy openpyxl
  python generate_dog_dataset.py

  # Train SVM on this dataset:
  python train_model.py --data canine_dog_dataset.csv --output cooling_svm_model.joblib
"""

import numpy as np
import pandas as pd
from datetime import datetime, timedelta

# -- Threshold constants (must match canine_model.h) -----------------------
BODY_TEMP_CUTOFF = 39.5   # C
RESP_RATE_LOWER  = 29.0   # cycles/min
AMBIENT_CUTOFF   = 30.0   # C

PHASE_MINS = 12            # minutes of monitoring per activity phase
SEED       = 42
rng        = np.random.default_rng(SEED)

# -- Dog roster: (dog_id, name, weight_kg, ambient_condition) --------------
# All dogs are 10-27 kg (vest-compatible medium size).
# "cool" ambient ~ 25 C (shaded/indoor)  -> ambient NOT stressed
# "warm" ambient ~ 31 C (Philippine heat) -> ambient IS stressed
DOGS = [
    ("D01", "Dog 1",  10, "warm"),
    ("D02", "Dog 2",  11, "cool"),
    ("D03", "Dog 3",  12, "warm"),
    ("D04", "Dog 4",  13, "cool"),
    ("D05", "Dog 5",  15, "warm"),
    ("D06", "Dog 6",  17, "cool"),
    ("D07", "Dog 7",  18, "warm"),
    ("D08", "Dog 8",  20, "warm"),
    ("D09", "Dog 9",  22, "cool"),
    ("D10", "Dog 10", 24, "warm"),
    ("D11", "Dog 11", 25, "cool"),
    ("D12", "Dog 12", 27, "warm"),
]

# Ambient temperature profiles (mean C, std dev)
AMBIENT = {
    "cool": (25.5, 0.5),   # controlled/shaded -- below 30 C threshold
    "warm": (31.2, 0.6),   # Philippine outdoor midday -- above 30 C threshold
}

# Session start times (12 dogs over 2 days, 1-hour slots each)
SESSION_STARTS = {
    "D01": datetime(2026, 6,  9,  8, 0),
    "D02": datetime(2026, 6,  9,  9, 0),
    "D03": datetime(2026, 6,  9, 10, 0),
    "D04": datetime(2026, 6,  9, 11, 0),
    "D05": datetime(2026, 6,  9, 13, 0),
    "D06": datetime(2026, 6,  9, 14, 0),
    "D07": datetime(2026, 6,  9, 15, 0),
    "D08": datetime(2026, 6,  9, 16, 0),
    "D09": datetime(2026, 6, 10,  8, 0),
    "D10": datetime(2026, 6, 10,  9, 0),
    "D11": datetime(2026, 6, 10, 10, 0),
    "D12": datetime(2026, 6, 10, 11, 0),
}


def get_params(weight_kg: float) -> dict:
    """
    Return size-specific physiological parameters interpolated by weight
    within the 10-27 kg range.
    w = 0.0 at 10 kg (lightest), 1.0 at 27 kg (heaviest).
    """
    w = (weight_kg - 10.0) / 17.0          # 0.0 -> 1.0
    w = float(np.clip(w, 0.0, 1.0))
    return dict(
        body_rest     = 38.70 - w * 0.28,  # 38.70 C (10 kg) -> 38.42 C (27 kg)
        body_noise    = 0.10,
        walk_body_max = 0.35  - w * 0.12,  # max temp rise during WALK
        run_body_max  = 0.88  - w * 0.25,  # additional temp rise during RUN
        resp_rest     = 22.0  - w * 7.0,   # 22 bpm (10 kg) -> 15 bpm (27 kg)
        resp_noise    = 2.0   - w * 0.5,
        walk_resp_max = 13.0  - w * 4.0,   # max resp rise during WALK
        run_resp_max  = 26.0  - w * 8.0,   # additional resp rise during RUN
    )


def _linear_rise(base: float, rise: float, minute: int, noise_std: float) -> float:
    """Value rises linearly from base (minute 0) to base+rise (minute 11) with noise."""
    progress = minute / max(PHASE_MINS - 1, 1)
    return base + rise * progress + rng.normal(0, noise_std)


def generate_dog_session(dog_id, name, weight_kg, ambient_cond):
    p              = get_params(weight_kg)
    amb_mean, amb_std = AMBIENT[ambient_cond]
    start          = SESSION_STARTS[dog_id]
    records        = []

    for phase_idx, phase in enumerate(["REST", "WALK", "RUN"]):
        phase_start = start + timedelta(minutes=phase_idx * PHASE_MINS)

        for minute in range(PHASE_MINS):
            ts = phase_start + timedelta(minutes=minute)

            # -- Body temperature ------------------------------------------
            if phase == "REST":
                body_temp = p["body_rest"] + rng.normal(0, p["body_noise"])
            elif phase == "WALK":
                body_temp = _linear_rise(
                    p["body_rest"], p["walk_body_max"],
                    minute, p["body_noise"] * 1.2
                )
            else:  # RUN starts where WALK peak left off
                walk_peak = p["body_rest"] + p["walk_body_max"]
                body_temp = _linear_rise(
                    walk_peak, p["run_body_max"],
                    minute, p["body_noise"] * 1.5
                )

            # -- Respiratory rate ------------------------------------------
            if phase == "REST":
                resp_rate = p["resp_rest"] + rng.normal(0, p["resp_noise"])
            elif phase == "WALK":
                resp_rate = _linear_rise(
                    p["resp_rest"], p["walk_resp_max"],
                    minute, p["resp_noise"] * 1.2
                )
            else:  # RUN
                walk_peak_r = p["resp_rest"] + p["walk_resp_max"]
                resp_rate   = _linear_rise(
                    walk_peak_r, p["run_resp_max"],
                    minute, p["resp_noise"] * 1.5
                )

            # -- Ambient temperature ---------------------------------------
            ambient_temp = rng.normal(amb_mean, amb_std)

            # -- Clamp to physiological bounds ----------------------------
            body_temp    = float(np.clip(body_temp,    37.5, 42.0))
            resp_rate    = float(np.clip(resp_rate,     5.0, 80.0))
            ambient_temp = float(np.clip(ambient_temp, 18.0, 40.0))

            # -- 2-of-3 labeling ------------------------------------------
            bs  = int(body_temp    >  BODY_TEMP_CUTOFF)
            rs  = int(resp_rate    >= RESP_RATE_LOWER)
            ams = int(ambient_temp >  AMBIENT_CUTOFF)
            sc  = bs + rs + ams
            label = int(sc >= 2)

            records.append({
                "timestamp":         ts.strftime("%Y-%m-%d %H:%M:%S"),
                "dog_id":            dog_id,
                "dog_name":          name,
                "weight_kg":         weight_kg,
                "ambient_condition": ambient_cond,
                "activity_phase":    phase,
                "elapsed_min":       minute,
                "body_temp":         round(body_temp,    2),
                "respiratory_rate":  round(resp_rate,    1),
                "ambient_temp":      round(ambient_temp, 2),
                "body_stressed":     bs,
                "resp_stressed":     rs,
                "ambient_stressed":  ams,
                "stressed_count":    sc,
                "label":             label,
            })

    return records


# -- Excel export -----------------------------------------------------------

def save_excel(df: pd.DataFrame, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY     = "1F3864"
    WHITE    = "FFFFFF"
    GREEN    = "C6EFCE"    # NORMAL rows
    RED      = "FFC7CE"    # HEAT_STRESSED rows
    YELLOW   = "FFFFC0"    # partial stress (summary sheet)

    PHASE_FILL = {
        "REST": PatternFill("solid", fgColor="BDD7EE"),   # blue
        "WALK": PatternFill("solid", fgColor="FFE699"),   # yellow
        "RUN":  PatternFill("solid", fgColor="F4CCCC"),   # pink-red
    }

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True, size=10)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def write_header(ws, columns, row=1):
        for ci, name in enumerate(columns, 1):
            c = ws.cell(row=row, column=ci, value=name)
            c.fill, c.font, c.alignment, c.border = header_fill, header_font, center, thin
        ws.row_dimensions[row].height = 18

    def auto_width(ws, columns, extra=4):
        for ci, col_name in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col_name)), 10) + extra

    wb = Workbook()

    # -- Sheet 1: Raw Data -------------------------------------------------
    ws1 = wb.active
    ws1.title = "Raw Data"
    ws1.freeze_panes = "A2"
    columns = list(df.columns)
    write_header(ws1, columns)
    phase_col_idx = columns.index("activity_phase") + 1

    for er, (_, row) in enumerate(df.iterrows(), start=2):
        row_fill = (PatternFill("solid", fgColor=RED)
                    if row["label"] == 1
                    else PatternFill("solid", fgColor=GREEN))
        for ci, value in enumerate(row, start=1):
            c = ws1.cell(row=er, column=ci, value=value)
            c.alignment = center
            c.border    = thin
            c.fill = (PHASE_FILL.get(row["activity_phase"], row_fill)
                      if ci == phase_col_idx else row_fill)
    auto_width(ws1, columns)

    # -- Sheet 2: Dog Summary ----------------------------------------------
    ws2 = wb.create_sheet("Dog Summary")
    ws2.freeze_panes = "A2"
    summary = (
        df.groupby(["dog_id", "dog_name", "weight_kg",
                    "ambient_condition", "activity_phase"], sort=False)
        .agg(
            avg_body_temp   =("body_temp",        "mean"),
            max_body_temp   =("body_temp",        "max"),
            avg_resp_rate   =("respiratory_rate", "mean"),
            max_resp_rate   =("respiratory_rate", "max"),
            avg_ambient_temp=("ambient_temp",     "mean"),
            heat_stressed   =("label",            "sum"),
            total_readings  =("label",            "count"),
        )
        .round(2).reset_index()
    )
    summary["pct_stressed_%"] = (
        summary["heat_stressed"] / summary["total_readings"] * 100
    ).round(1)

    s_cols = list(summary.columns)
    write_header(ws2, s_cols)
    for er, (_, row) in enumerate(summary.iterrows(), start=2):
        pct = row["pct_stressed_%"]
        fill = (PatternFill("solid", fgColor=RED)   if pct >= 50 else
                PatternFill("solid", fgColor=YELLOW) if pct > 0  else
                PatternFill("solid", fgColor=GREEN))
        for ci, value in enumerate(row, start=1):
            c = ws2.cell(row=er, column=ci, value=value)
            c.fill, c.alignment, c.border = fill, center, thin
    auto_width(ws2, s_cols, extra=5)

    # -- Sheet 3: Phase Overview -------------------------------------------
    ws3 = wb.create_sheet("Phase Overview")
    ws3.freeze_panes = "A2"
    phase_ov = (
        df.groupby("activity_phase")
        .agg(
            avg_body_temp    =("body_temp",        "mean"),
            max_body_temp    =("body_temp",        "max"),
            avg_resp_rate    =("respiratory_rate", "mean"),
            max_resp_rate    =("respiratory_rate", "max"),
            avg_ambient      =("ambient_temp",     "mean"),
            pct_heat_stressed=("label",            "mean"),
            total_readings   =("label",            "count"),
        )
        .round(3).reset_index()
    )
    phase_ov["pct_heat_stressed"] = (phase_ov["pct_heat_stressed"] * 100).round(1)
    p_cols = list(phase_ov.columns)
    write_header(ws3, p_cols)
    for er, (_, row) in enumerate(phase_ov.iterrows(), start=2):
        fill = PHASE_FILL.get(row["activity_phase"], PatternFill("solid", fgColor=GREEN))
        for ci, value in enumerate(row, start=1):
            c = ws3.cell(row=er, column=ci, value=value)
            c.fill, c.alignment, c.border = fill, center, thin
    auto_width(ws3, p_cols, extra=5)

    wb.save(path)


# -- Main ------------------------------------------------------------------

def main():
    all_records = []
    for dog_id, name, weight_kg, ambient in DOGS:
        records = generate_dog_session(dog_id, name, weight_kg, ambient)
        all_records.extend(records)

    df = pd.DataFrame(all_records)

    csv_path  = "canine_dog_dataset.csv"
    xlsx_path = "canine_dog_dataset.xlsx"

    df.to_csv(csv_path, index=False)
    print(f"[CSV]  Saved {len(df)} rows to '{csv_path}'")

    save_excel(df, xlsx_path)
    print(f"[XLSX] Saved workbook to '{xlsx_path}'")

    print(f"\nDataset shape  : {df.shape[0]} rows x {df.shape[1]} columns")
    print(f"Dogs monitored : {df['dog_id'].nunique()}")
    print(f"Weight range   : {df['weight_kg'].min()} - {df['weight_kg'].max()} kg")

    print("\n--- Label distribution ---")
    ld = df["label"].value_counts().rename({0: "NORMAL (0)", 1: "HEAT_STRESSED (1)"})
    print(ld.to_string())

    print("\n--- Average readings by activity phase ---")
    phase_stats = (
        df.groupby("activity_phase")
        .agg(
            avg_body_temp  =("body_temp",        "mean"),
            avg_resp_rate  =("respiratory_rate", "mean"),
            avg_ambient    =("ambient_temp",     "mean"),
            pct_stressed   =("label",            "mean"),
        )
        .round(2)
    )
    phase_stats["pct_stressed"] = (phase_stats["pct_stressed"] * 100).round(1)
    print(phase_stats.to_string())

    print("\n--- Average readings by weight group ---")
    df["weight_group"] = pd.cut(df["weight_kg"],
                                bins=[9, 14, 20, 27],
                                labels=["10-14 kg", "15-20 kg", "21-27 kg"])
    wt_stats = (
        df.groupby("weight_group", observed=True)
        .agg(
            avg_body_temp  =("body_temp",        "mean"),
            avg_resp_rate  =("respiratory_rate", "mean"),
            pct_stressed   =("label",            "mean"),
        )
        .round(2)
    )
    wt_stats["pct_stressed"] = (wt_stats["pct_stressed"] * 100).round(1)
    print(wt_stats.to_string())

    print(f"\nNext step:")
    print(f"  python train_model.py --data {csv_path} --output cooling_svm_model.joblib")


if __name__ == "__main__":
    main()
